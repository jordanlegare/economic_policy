#!/usr/bin/env python3
"""Build the provisional U.S. 20-sector I-O proxy from EPA USEEIO v2.5.

This is a reproducible fallback for the U.S. production network while the
current-vintage BEA annual Input-Output artifact is pending. It must not be
described as the final/current BEA calibration.

Source
------
EPA USEEIO v2.5 detailed model ``USEEIOv2.5-catbird-22.xlsx``, DOI
10.23719/1532178.  The workbook supplies:

* ``A_d``: domestic direct-requirements matrix, upstream commodity by
  downstream commodity;
* ``q``: detailed commodity output used for downstream aggregation weights;
* ``commodities_meta``: detailed BEA/USEEIO commodity codes and metadata.

EPA documents that the detailed v2.5 models published for 2022 use underlying
U.S. input-output data for 2017.  The 2022 suffix refers to the model vintage
and coupled-import treatment, not a 2022 detailed domestic I-O table.

Aggregation
-----------
The workbook contains 402 detailed commodities. Four accounting/adjustment
commodities (scrap, used/secondhand goods, noncomparable imports, and the
rest-of-world adjustment) are excluded, leaving 398 commodities mapped into
the simulator's 20 sectors.

For downstream model sector J and upstream model sector I:

    A_proxy[J][I] =
        sum_{j in J} q_j * sum_{i in I} max(A_d[i][j], 0)
        / sum_{j in J} q_j

This preserves USEEIO's domestic direct-requirements orientation and uses
commodity-output weights. Negative accounting/redefinition adjustments are
not propagated, matching the repository's BEA-builder treatment of negative
direct-requirement cells.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from pathlib import Path

MODEL_CODES = [
    "11", "21", "22", "23", "31-33", "42", "44-45", "48-49", "51", "52",
    "53", "54", "55", "56", "61", "62", "71", "72", "81", "91",
]
MODEL_NAMES = [
    "Agriculture, forestry, fishing & hunting",
    "Mining, quarrying, oil & gas",
    "Utilities",
    "Construction",
    "Manufacturing",
    "Wholesale trade",
    "Retail trade",
    "Transportation & warehousing",
    "Information & cultural industries",
    "Finance & insurance",
    "Real estate, rental & leasing",
    "Professional, scientific & technical services",
    "Management of companies & enterprises",
    "Administrative, support & waste services",
    "Educational services",
    "Health care & social assistance",
    "Arts, entertainment & recreation",
    "Accommodation & food services",
    "Other services (except public administration)",
    "Public administration",
]
EXCLUDED_CODES = {
    "S00401": "Scrap",
    "S00402": "Used and secondhand goods",
    "S00300": "Noncomparable imports",
    "S00900": "Rest of the world adjustment",
}
SOURCE_URL = (
    "https://pasteur.epa.gov/uploads/10.23719/1532178/"
    "USEEIOv2.5-catbird-22.xlsx"
)
SOURCE_DOI = "10.23719/1532178"
SOURCE_DATASET = "USEEIO v2.5 Models"
SOURCE_MODEL = "USEEIOv2.5-catbird-22"
DOMESTIC_IO_BASIS_YEAR = 2017
MODEL_REFERENCE_YEAR = 2022


def normalize_code(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).upper()
    return str(value).strip().upper()


def model_sector(code: object) -> int | None:
    raw = re.sub(r"[^A-Z0-9]", "", normalize_code(code))
    if raw in EXCLUDED_CODES:
        return None

    # BEA detailed presentation codes used by current USEEIO workbooks.
    if raw in {"4A0000", "4B0000"}:
        return 6
    if raw.startswith(("S001", "S002", "S005", "S006", "S007", "GSL")):
        return 19

    digits = "".join(ch for ch in raw if ch.isdigit())
    if len(digits) < 2:
        return None
    prefix = digits[:2]
    if prefix in {"31", "32", "33"}:
        return 4
    if prefix in {"44", "45"}:
        return 6
    if prefix in {"48", "49"}:
        return 7

    direct = {
        "11": 0, "21": 1, "22": 2, "23": 3, "42": 5,
        "51": 8, "52": 9, "53": 10, "54": 11, "55": 12, "56": 13,
        "61": 14, "62": 15, "71": 16, "72": 17, "81": 18,
        "91": 19, "92": 19,
    }
    return direct.get(prefix)


def aggregate_matrix(
    detail_codes: list[str],
    direct_domestic: list[list[float]],
    commodity_output: list[float],
) -> tuple[list[list[float]], list[int], int]:
    n = len(detail_codes)
    if (
        n == 0
        or len(direct_domestic) != n
        or any(len(row) != n for row in direct_domestic)
        or len(commodity_output) != n
    ):
        raise ValueError("USEEIO A_d, q, and detail labels must share one square dimension")

    mapped = [model_sector(code) for code in detail_codes]
    counts = [0] * len(MODEL_CODES)
    denominators = [0.0] * len(MODEL_CODES)
    numerator = [[0.0] * len(MODEL_CODES) for _ in MODEL_CODES]
    mapped_positive_cells = 0

    # A_d orientation is [upstream commodity][downstream commodity].
    for downstream_detail, downstream_sector in enumerate(mapped):
        if downstream_sector is None:
            continue
        weight = float(commodity_output[downstream_detail])
        if weight <= 0.0:
            raise ValueError(
                f"Mapped commodity {detail_codes[downstream_detail]} has non-positive q"
            )
        counts[downstream_sector] += 1
        denominators[downstream_sector] += weight

        for upstream_detail, upstream_sector in enumerate(mapped):
            if upstream_sector is None:
                continue
            coeff = float(direct_domestic[upstream_detail][downstream_detail])
            if coeff <= 0.0:
                continue
            numerator[downstream_sector][upstream_sector] += weight * coeff
            mapped_positive_cells += 1

    missing = [
        MODEL_CODES[i] for i, denominator in enumerate(denominators)
        if denominator <= 0.0
    ]
    if missing:
        raise ValueError(
            "No positive USEEIO commodity-output denominator for model sectors: "
            + ", ".join(missing)
        )

    out = [
        [
            numerator[downstream][upstream] / denominators[downstream]
            for upstream in range(len(MODEL_CODES))
        ]
        for downstream in range(len(MODEL_CODES))
    ]
    for downstream, row in enumerate(out):
        total = sum(row)
        if not 0.0 < total < 1.0:
            raise ValueError(
                f"Implausible proxy row sum for {MODEL_CODES[downstream]}: {total:.12f}"
            )

    return out, counts, mapped_positive_cells


def load_useeio(
    path: Path,
) -> tuple[list[str], list[list[float]], list[float]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit(
            "openpyxl is required to read the official USEEIO workbook"
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    required = {"A_d", "q", "commodities_meta"}
    missing_sheets = required.difference(workbook.sheetnames)
    if missing_sheets:
        raise SystemExit(
            "Workbook is missing USEEIO sheets: " + ", ".join(sorted(missing_sheets))
        )

    meta_sheet = workbook["commodities_meta"]
    detail_codes: list[str] = []
    detail_ids: list[str] = []
    for row in meta_sheet.iter_rows(min_row=2, values_only=True):
        # Current USEEIO model format: Index, ID, Name, Code, Location, ...
        detail_id = str(row[1] or "").strip()
        code = normalize_code(row[3])
        if code:
            detail_ids.append(detail_id)
            detail_codes.append(code)

    def label_code(label: object) -> str:
        return str(label or "").split("/", 1)[0].strip().upper()

    a_sheet = workbook["A_d"]
    header = next(a_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    column_labels = [str(value or "") for value in header[1:]]
    a_rows = list(a_sheet.iter_rows(min_row=2, values_only=True))
    if len(detail_codes) != len(column_labels) or len(a_rows) != len(detail_codes):
        raise SystemExit("USEEIO A_d dimensions do not match commodities_meta")

    direct_domestic: list[list[float]] = []
    for idx, row in enumerate(a_rows):
        expected = detail_codes[idx]
        if label_code(column_labels[idx]) != expected or label_code(row[0]) != expected:
            raise SystemExit(
                f"USEEIO A_d commodity-order mismatch at {idx}: expected {expected}"
            )
        direct_domestic.append([float(value or 0.0) for value in row[1:]])

    q_rows = list(workbook["q"].iter_rows(min_row=2, values_only=True))
    if len(q_rows) != len(detail_codes):
        raise SystemExit("USEEIO q dimensions do not match commodities_meta")
    commodity_output: list[float] = []
    for idx, row in enumerate(q_rows):
        expected = detail_codes[idx]
        if label_code(row[0]) != expected:
            raise SystemExit(
                f"USEEIO q commodity-order mismatch at {idx}: expected {expected}"
            )
        commodity_output.append(float(row[1] or 0.0))

    return detail_codes, direct_domestic, commodity_output


def render_csv(matrix: list[list[float]]) -> str:
    lines = ["downstream_code,downstream_name," + ",".join(MODEL_CODES)]
    for code, name, row in zip(MODEL_CODES, MODEL_NAMES, matrix):
        quoted = '"' + name.replace('"', '""') + '"'
        lines.append(",".join([code, quoted] + [f"{value:.12f}" for value in row]))
    return "\n".join(lines) + "\n"


def render_header(matrix: list[list[float]]) -> str:
    rows = [
        "    {{" + ", ".join(f"{value:.12f}" for value in row) + "}}"
        for row in matrix
    ]
    maximum = max(sum(row) for row in matrix)
    return (
        "#pragma once\n\n"
        '#include "trade_network.hpp"\n\n'
        "namespace cad::generated {\n\n"
        "// Provisional U.S. production-network proxy generated from EPA USEEIO v2.5.\n"
        "// Model: USEEIOv2.5-catbird-22; underlying detailed U.S. I-O basis: 2017.\n"
        "// Source matrix: A_d (domestic direct requirements); weighted by q commodity output.\n"
        "// Orientation: [downstream][upstream]. Do not edit by hand.\n"
        "inline constexpr TradeInputOutputMatrix kEpaUseeioUsProxyMatrix{{\n"
        + ",\n".join(rows)
        + "\n}};\n"
        f"inline constexpr double kEpaUseeioUsProxyMaximumDomesticIntermediateShare = {maximum:.12f};\n\n"
        "}  // namespace cad::generated\n"
    )


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_or_verify(path: Path, content: str, verify: bool) -> None:
    if verify:
        if not path.exists() or path.read_text(encoding="utf-8") != content:
            raise RuntimeError(f"Generated artifact drift: {path}")
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--input",
        required=True,
        help="Official EPA USEEIOv2.5-catbird-22 XLSX",
    )
    parser.add_argument("--csv", default="data/calibration/useeio_us_proxy_matrix.csv")
    parser.add_argument("--header", default="include/generated/trade_io_us_proxy.hpp")
    parser.add_argument(
        "--provenance",
        default="data/calibration/useeio_us_proxy_provenance.json",
    )
    parser.add_argument("--source-url", default=SOURCE_URL)
    parser.add_argument("--verify", action="store_true")
    args = parser.parse_args()

    source = Path(args.input)
    if not source.is_file():
        raise SystemExit("--input must point to the official EPA USEEIO v2.5 workbook")

    detail_codes, direct_domestic, commodity_output = load_useeio(source)
    matrix, counts, mapped_positive_cells = aggregate_matrix(
        detail_codes, direct_domestic, commodity_output
    )
    csv_text = render_csv(matrix)
    header_text = render_header(matrix)

    mapped_commodities = sum(counts)
    excluded = sorted(code for code in detail_codes if code in EXCLUDED_CODES)
    provenance = {
        "agency": "U.S. Environmental Protection Agency, Office of Research and Development",
        "dataset": SOURCE_DATASET,
        "model": SOURCE_MODEL,
        "doi": SOURCE_DOI,
        "source_url": args.source_url,
        "source_file": source.name,
        "source_sha256": sha256(source),
        "source_matrix": "A_d",
        "source_matrix_orientation": "upstream_commodity_by_downstream_commodity",
        "weight_vector": "q",
        "source_detail_commodities": len(detail_codes),
        "mapped_detail_commodities": mapped_commodities,
        "excluded_adjustment_codes": excluded,
        "model_sectors": len(MODEL_CODES),
        "sector_detail_counts": dict(zip(MODEL_CODES, counts)),
        "mapped_positive_cells": mapped_positive_cells,
        "aggregation": (
            "commodity-output-weighted aggregation of positive domestic direct "
            "requirements from USEEIO A_d into 20 downstream-by-upstream sectors"
        ),
        "negative_adjustments": "clipped_to_zero",
        "domestic_io_basis_year": DOMESTIC_IO_BASIS_YEAR,
        "model_reference_year": MODEL_REFERENCE_YEAR,
        "activation_status": "provisional_us_proxy_pending_current_bea_artifact",
        "empirical_us_current_vintage": False,
        "replacement_contract": (
            "include/generated/trade_io_us_bea.hpp plus an independently reviewed "
            "include/generated/trade_io_us_bea_certified.hpp marker supersede this proxy"
        ),
        "maximum_domestic_intermediate_share": max(sum(row) for row in matrix),
    }
    provenance_text = json.dumps(provenance, indent=2, sort_keys=True) + "\n"

    write_or_verify(Path(args.csv), csv_text, args.verify)
    write_or_verify(Path(args.header), header_text, args.verify)
    write_or_verify(Path(args.provenance), provenance_text, args.verify)
    print(provenance_text, end="")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
